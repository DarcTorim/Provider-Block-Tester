#!/usr/bin/env python3
"""
╔═══════════════════════════════════════════════════════════╗
║                                                           ║
║        🛡️  PROVIDER BLOCK TESTER  v3.1                    ║
║                                                           ║
║     Тестер блокировок с Excel отчётом на рабочем столе    ║
║                                                           ║
║     Установка: pip install requests dnspython colorama openpyxl ║
║     Запуск: python provider_tester.py --excel             ║
║                                                           ║
╚═══════════════════════════════════════════════════════════╝
"""

# ============================================================================
# ЗАВИСИМОСТИ
# ============================================================================
import socket
import dns.resolver
import requests
import time
import ssl
import json
import os
import sys
import argparse
from datetime import datetime
from typing import Dict, List, Optional, Tuple
from dataclasses import dataclass, field

try:
    from colorama import init, Fore, Style

    init(autoreset=True)
except ImportError:
    print("⚠️  Установите colorama: pip install colorama")


    class DummyColor:
        def __getattr__(self, name): return ""


    Fore = Style = DummyColor()


    def init(**kwargs):
        pass

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️  Установите openpyxl для Excel отчётов: pip install openpyxl")


# ============================================================================
# КОНФИГУРАЦИЯ
# ============================================================================

@dataclass
class TestResource:
    """Модель ресурса для тестирования"""
    name: str
    domain: str
    url: str
    category: str


@dataclass
class Config:
    """Настройки тестера"""
    TEST_RESOURCES: List[TestResource] = field(default_factory=list)
    PUBLIC_DNS_SERVERS: List[str] = field(default_factory=list)
    DNS_TIMEOUT: float = 3.0
    HTTP_TIMEOUT: float = 5.0
    TCP_TIMEOUT: float = 3.0
    TEST_PORTS: List[int] = field(default_factory=list)
    REPORT_DIR: str = "reports"
    EXCEL_ON_DESKTOP: bool = True  # Сохранять Excel на рабочий стол

    def __post_init__(self):
        if not self.TEST_RESOURCES:
            self.TEST_RESOURCES = [
                TestResource("Rutracker", "rutracker.org", "https://rutracker.org", "Torrent"),
                TestResource("Instagram", "instagram.com", "https://instagram.com", "Соцсети"),
                TestResource("Facebook", "facebook.com", "https://facebook.com", "Соцсети"),
                TestResource("Twitter/X", "twitter.com", "https://twitter.com", "Соцсети"),
                TestResource("Discord", "discord.com", "https://discord.com", "Мессенджеры"),
                TestResource("GitHub", "github.com", "https://github.com", "IT"),
                TestResource("Telegram Web", "web.telegram.org", "https://web.telegram.org", "Мессенджеры"),
                TestResource("YouTube", "youtube.com", "https://youtube.com", "Видео"),
                TestResource("Netflix", "netflix.com", "https://netflix.com", "Видео"),
                TestResource("Wikipedia", "wikipedia.org", "https://wikipedia.org", "Новости"),
                TestResource("Reddit", "reddit.com", "https://reddit.com", "Соцсети"),
                TestResource("Twitch", "twitch.tv", "https://twitch.tv", "Видео"),
            ]

        if not self.PUBLIC_DNS_SERVERS:
            self.PUBLIC_DNS_SERVERS = [
                "8.8.8.8",  # Google
                "1.1.1.1",  # Cloudflare
                "9.9.9.9",  # Quad9
                "77.88.8.8",  # Yandex
            ]

        if not self.TEST_PORTS:
            self.TEST_PORTS = [80, 443, 8080, 8443]


DEFAULT_CONFIG = Config()


# ============================================================================
# МОДЕЛИ РЕЗУЛЬТАТОВ
# ============================================================================

@dataclass
class DNSResult:
    provider_ip: Optional[str]
    public_ip: Optional[str]
    is_blocked: bool
    is_spoofed: bool
    response_time: float
    error: Optional[str] = None


@dataclass
class HTTPResult:
    is_accessible: bool
    status_code: Optional[int]
    response_time: float
    error: Optional[str] = None
    ssl_valid: bool = True


@dataclass
class TCPResult:
    port: int
    is_open: bool
    response_time: float
    error: Optional[str] = None


@dataclass
class ResourceTestResult:
    resource: TestResource
    dns_result: DNSResult
    http_result: HTTPResult
    tcp_results: List[TCPResult]
    block_type: str
    confidence: float
    timestamp: str


# ============================================================================
# ТЕСТЕР
# ============================================================================

class ProviderTester:
    def __init__(self, config: Config = None):
        self.config = config or Config()
        self.results: List[ResourceTestResult] = []

    def test_dns(self, domain: str) -> DNSResult:
        start_time = time.time()

        provider_ip = None
        try:
            resolver = dns.resolver.Resolver()
            resolver.timeout = self.config.DNS_TIMEOUT
            resolver.lifetime = self.config.DNS_TIMEOUT
            answers = resolver.resolve(domain, 'A')
            for rdata in answers:
                provider_ip = str(rdata)
                break
        except Exception as e:
            pass

        public_ip = None
        try:
            resolver = dns.resolver.Resolver()
            resolver.nameservers = [self.config.PUBLIC_DNS_SERVERS[0]]
            resolver.timeout = self.config.DNS_TIMEOUT
            resolver.lifetime = self.config.DNS_TIMEOUT
            answers = resolver.resolve(domain, 'A')
            for rdata in answers:
                public_ip = str(rdata)
                break
        except Exception as e:
            pass

        response_time = time.time() - start_time

        is_blocked = False
        is_spoofed = False
        error = None

        if provider_ip is None:
            is_blocked = True
            error = "DNS не ответил"
        elif provider_ip in ["0.0.0.0", "127.0.0.1"]:
            is_blocked = True
            error = "DNS заглушка"
        elif public_ip and provider_ip != public_ip:
            is_spoofed = True
            error = "IP не совпадает с публичным"

        return DNSResult(
            provider_ip=provider_ip,
            public_ip=public_ip,
            is_blocked=is_blocked,
            is_spoofed=is_spoofed,
            response_time=response_time,
            error=error
        )

    def test_http(self, url: str) -> HTTPResult:
        start_time = time.time()

        try:
            response = requests.get(
                url,
                timeout=self.config.HTTP_TIMEOUT,
                verify=True,
                allow_redirects=False
            )
            response_time = time.time() - start_time

            return HTTPResult(
                is_accessible=response.status_code < 500,
                status_code=response.status_code,
                response_time=response_time,
                ssl_valid=True
            )
        except requests.exceptions.SSLError as e:
            return HTTPResult(
                is_accessible=False,
                status_code=None,
                response_time=time.time() - start_time,
                error=f"SSL ошибка: {str(e)}",
                ssl_valid=False
            )
        except requests.exceptions.Timeout:
            return HTTPResult(
                is_accessible=False,
                status_code=None,
                response_time=time.time() - start_time,
                error="Timeout"
            )
        except requests.exceptions.ConnectionError as e:
            return HTTPResult(
                is_accessible=False,
                status_code=None,
                response_time=time.time() - start_time,
                error="Connection Error"
            )
        except Exception as e:
            return HTTPResult(
                is_accessible=False,
                status_code=None,
                response_time=time.time() - start_time,
                error=str(e)
            )

    def test_tcp_port(self, domain: str, port: int) -> TCPResult:
        start_time = time.time()

        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(self.config.TCP_TIMEOUT)
            result = sock.connect_ex((domain, port))
            response_time = time.time() - start_time
            sock.close()

            return TCPResult(
                port=port,
                is_open=(result == 0),
                response_time=response_time
            )
        except Exception as e:
            return TCPResult(
                port=port,
                is_open=False,
                response_time=time.time() - start_time,
                error=str(e)
            )

    def determine_block_type(self, dns: DNSResult, http: HTTPResult, tcp: List[TCPResult]) -> Tuple[str, float]:
        if not dns.is_blocked and not dns.is_spoofed and http.is_accessible:
            return "none", 1.0

        if dns.is_blocked:
            confidence = 0.9
            if http.is_accessible:
                confidence = 0.7
            return "dns", confidence

        if dns.is_spoofed:
            if not http.is_accessible:
                return "dns_spoof", 0.85
            else:
                return "dns_spoof_partial", 0.6

        if not http.is_accessible:
            if "Timeout" in (http.error or ""):
                tcp_open = any(t.is_open for t in tcp)
                if tcp_open:
                    return "http_filter", 0.8
                else:
                    return "tcp_block", 0.75
            elif "Connection" in (http.error or ""):
                return "tcp_reset", 0.8
            else:
                return "unknown", 0.5

        return "none", 1.0

    def test_resource(self, resource: TestResource) -> ResourceTestResult:
        print(f"  📍 Тестируем {resource.name}...", end=" ", flush=True)

        dns_result = self.test_dns(resource.domain)
        http_result = self.test_http(resource.url)

        tcp_results = []
        for port in self.config.TEST_PORTS[:2]:
            tcp_results.append(self.test_tcp_port(resource.domain, port))

        block_type, confidence = self.determine_block_type(dns_result, http_result, tcp_results)

        print("✅" if block_type == "none" else "⚠️")

        return ResourceTestResult(
            resource=resource,
            dns_result=dns_result,
            http_result=http_result,
            tcp_results=tcp_results,
            block_type=block_type,
            confidence=confidence,
            timestamp=datetime.now().isoformat()
        )

    def run_all_tests(self) -> List[ResourceTestResult]:
        print("\n" + "=" * 70)
        print("🚀 ЗАПУСК ПОЛНОГО ТЕСТИРОВАНИЯ")
        print("=" * 70 + "\n")

        self.results = []

        for resource in self.config.TEST_RESOURCES:
            result = self.test_resource(resource)
            self.results.append(result)

        return self.results

    def get_summary(self) -> Dict:
        total = len(self.results)
        blocked = sum(1 for r in self.results if r.block_type != "none")
        dns_blocked = sum(1 for r in self.results if "dns" in r.block_type)
        http_blocked = sum(1 for r in self.results if r.block_type in ["http_filter", "tcp_reset", "tcp_block"])

        return {
            "total": total,
            "blocked": blocked,
            "accessible": total - blocked,
            "dns_blocks": dns_blocked,
            "http_blocks": http_blocked,
            "block_percentage": (blocked / total * 100) if total > 0 else 0
        }


# ============================================================================
# ГЕНЕРАТОР ОТЧЁТОВ
# ============================================================================

class ReportGenerator:
    def __init__(self, tester: ProviderTester):
        self.tester = tester
        self.report_dir = tester.config.REPORT_DIR
        os.makedirs(self.report_dir, exist_ok=True)

    def _get_desktop_path(self) -> str:
        """Получает путь к рабочему столу"""
        try:
            import winreg
            with winreg.OpenKey(winreg.HKEY_CURRENT_USER,
                                r"Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders") as key:
                desktop = winreg.QueryValueEx(key, "Desktop")[0]
                return desktop
        except:
            # Если не Windows или ошибка, используем стандартный путь
            return os.path.join(os.path.expanduser("~"), "Desktop")

    def generate_console_report(self):
        summary = self.tester.get_summary()

        print("\n" + "=" * 70)
        print(f"{Fore.CYAN}📊 СВОДНЫЙ ОТЧЁТ{Style.RESET_ALL}")
        print("=" * 70)

        print(f"\n📈 Общая статистика:")
        print(f"  Всего проверено: {Fore.WHITE}{summary['total']}{Style.RESET_ALL}")
        print(f"  Доступно: {Fore.GREEN}{summary['accessible']}{Style.RESET_ALL}")
        print(f"  Заблокировано: {Fore.RED}{summary['blocked']}{Style.RESET_ALL}")
        print(f"  Процент блокировок: {Fore.YELLOW}{summary['block_percentage']:.1f}%{Style.RESET_ALL}")

        print(f"\n🔍 Типы блокировок:")
        print(f"  DNS блокировки: {Fore.RED}{summary['dns_blocks']}{Style.RESET_ALL}")
        print(f"  HTTP/TCP блокировки: {Fore.RED}{summary['http_blocks']}{Style.RESET_ALL}")

        print(f"\n{'=' * 70}")
        print(f"{Fore.CYAN}📋 ДЕТАЛЬНЫЙ ОТЧЁТ{Style.RESET_ALL}")
        print(f"{'=' * 70}\n")

        for result in self.tester.results:
            status_icon = "✅" if result.block_type == "none" else "🚫"
            status_color = Fore.GREEN if result.block_type == "none" else Fore.RED

            print(f"{status_icon} {Fore.WHITE}{result.resource.name}{Style.RESET_ALL} ({result.resource.domain})")
            print(f"   Статус: {status_color}{self._get_block_description(result.block_type)}{Style.RESET_ALL}")
            print(f"   Уверенность: {Fore.YELLOW}{result.confidence * 100:.0f}%{Style.RESET_ALL}")

            if result.dns_result.provider_ip:
                dns_status = Fore.GREEN + "OK" if not result.dns_result.is_blocked else Fore.RED + "BLOCKED"
                print(f"   DNS: {dns_status} ({result.dns_result.provider_ip})")

            if result.http_result.status_code:
                http_color = Fore.GREEN if result.http_result.status_code < 400 else Fore.RED
                print(
                    f"   HTTP: {http_color}{result.http_result.status_code}{Style.RESET_ALL} ({result.http_result.response_time:.2f}s)")
            elif result.http_result.error:
                print(f"   HTTP: {Fore.RED}{result.http_result.error}{Style.RESET_ALL}")

            # Порты
            port_status = []
            for tcp in result.tcp_results:
                port_status.append(f"{tcp.port}: {'✅' if tcp.is_open else '❌'}")
            print(f"   Порты: {', '.join(port_status)}")

            if result.block_type != "none":
                print(f"   💡 Рекомендация: {self._get_recommendation(result.block_type)}")

            print()

        self._print_final_recommendations(summary)

    def _get_block_description(self, block_type: str) -> str:
        descriptions = {
            "none": "Доступен",
            "dns": "DNS блокировка",
            "dns_spoof": "DNS подмена",
            "dns_spoof_partial": "DNS подмена (частичная)",
            "http_filter": "HTTP фильтрация (DPI)",
            "tcp_reset": "TCP сброс соединения",
            "tcp_block": "TCP блокировка порта",
            "unknown": "Неизвестная блокировка"
        }
        return descriptions.get(block_type, block_type)

    def _get_recommendation(self, block_type: str) -> str:
        recommendations = {
            "dns": "Смените DNS на 1.1.1.1 или 8.8.8.8",
            "dns_spoof": "Используйте DoH/DoT или смените DNS",
            "dns_spoof_partial": "Попробуйте DNS over HTTPS",
            "http_filter": "Используйте GoodbyeDPI или VPN",
            "tcp_reset": "Нужен обход DPI или прокси",
            "tcp_block": "Проверьте альтернативные порты",
            "unknown": "Попробуйте VPN или Tor"
        }
        return recommendations.get(block_type, "Нет рекомендаций")

    def _get_block_reason(self, block_type: str) -> str:
        reasons = {
            "none": "Нет блокировки",
            "dns": "Провайдер не резолвит домен или возвращает заглушку",
            "dns_spoof": "Провайдер подменяет IP адрес на свой",
            "dns_spoof_partial": "Частичная подмена DNS",
            "http_filter": "DPI фильтрация HTTPS трафика",
            "tcp_reset": "Провайдер сбрасывает TCP соединение (RST)",
            "tcp_block": "Порт заблокирован на уровне фаервола",
            "unknown": "Причина не определена"
        }
        return reasons.get(block_type, "Неизвестно")

    def _print_final_recommendations(self, summary: Dict):
        print(f"\n{'=' * 70}")
        print(f"{Fore.CYAN}💡 ОБЩИЕ РЕКОМЕНДАЦИИ{Style.RESET_ALL}")
        print(f"{'=' * 70}\n")

        if summary['block_percentage'] == 0:
            print(f"{Fore.GREEN}✅ Ваш провайдер не блокирует проверенные ресурсы!{Style.RESET_ALL}")
        elif summary['block_percentage'] < 30:
            print(f"{Fore.YELLOW}⚠️  Частичные блокировки обнаружены{Style.RESET_ALL}")
            print("   → Рекомендуется использовать DNS 1.1.1.1")
            print("   → Для отдельных сайтов можно использовать прокси")
        elif summary['block_percentage'] < 70:
            print(f"{Fore.RED}🚫 Значительные блокировки обнаружены{Style.RESET_ALL}")
            print("   → Рекомендуется использовать средства обхода DPI")
            print("   → Рассмотрите VPN для важных ресурсов")
        else:
            print(f"{Fore.RED}🔴 Массовые блокировки обнаружены{Style.RESET_ALL}")
            print("   → VPN или прокси необходимы для доступа")
            print("   → Рассмотрите Tor для анонимности")

        print()

    def generate_excel_report(self, on_desktop: bool = True) -> str:
        """Генерирует Excel отчёт с полной информацией"""
        if not EXCEL_AVAILABLE:
            print(f"{Fore.RED}❌ openpyxl не установлен. Установите: pip install openpyxl{Style.RESET_ALL}")
            return None

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        # Определяем путь сохранения
        if on_desktop:
            save_path = self._get_desktop_path()
            filename = os.path.join(save_path, f"Отчёт_блокировки_{timestamp}.xlsx")
        else:
            os.makedirs(self.report_dir, exist_ok=True)
            filename = os.path.join(self.report_dir, f"report_{timestamp}.xlsx")

        wb = Workbook()
        ws = wb.active
        ws.title = "Результаты проверки"

        # Стили
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        block_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Заголовки
        headers = [
            "№",
            "Ресурс",
            "Домен",
            "Категория",
            "Статус",
            "Тип блокировки",
            "Причина блокировки",
            "DNS IP (Провайдер)",
            "DNS IP (Публичный)",
            "DNS Статус",
            "HTTP Код",
            "HTTP Статус",
            "Порт 80",
            "Порт 443",
            "Уверенность %",
            "Рекомендация",
            "Время проверки"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Данные
        for row_idx, result in enumerate(self.tester.results, 2):
            # Статус
            status = "✅ Доступен" if result.block_type == "none" else "🚫 Заблокирован"
            status_fill = ok_fill if result.block_type == "none" else block_fill

            # DNS статус
            if result.dns_result.is_blocked:
                dns_status = "❌ Заблокирован"
            elif result.dns_result.is_spoofed:
                dns_status = "⚠️ Подмена"
            else:
                dns_status = "✅ OK"

            # HTTP статус
            if result.http_result.status_code:
                http_status = "✅ Работает" if result.http_result.status_code < 400 else "❌ Ошибка"
            else:
                http_status = f"❌ {result.http_result.error}" if result.http_result.error else "❌ Не доступен"

            # Порты
            port_80 = "✅ Открыт" if any(t.port == 80 and t.is_open for t in result.tcp_results) else "❌ Закрыт"
            port_443 = "✅ Открыт" if any(t.port == 443 and t.is_open for t in result.tcp_results) else "❌ Закрыт"

            data = [
                row_idx - 1,
                result.resource.name,
                result.resource.domain,
                result.resource.category,
                status,
                self._get_block_description(result.block_type),
                self._get_block_reason(result.block_type),
                result.dns_result.provider_ip or "N/A",
                result.dns_result.public_ip or "N/A",
                dns_status,
                result.http_result.status_code or "N/A",
                http_status,
                port_80,
                port_443,
                f"{result.confidence * 100:.0f}%",
                self._get_recommendation(result.block_type),
                datetime.now().strftime("%d.%m.%Y %H:%M")
            ]

            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                # Цвет строки
                if col_idx == 5:  # Статус
                    if result.block_type == "none":
                        cell.fill = ok_fill
                    elif "dns" in result.block_type:
                        cell.fill = block_fill
                    else:
                        cell.fill = warning_fill

        # Автоширина колонок
        column_widths = [10, 20, 25, 15, 15, 25, 40, 20, 20, 15, 12, 25, 15, 15, 15, 35, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Закрепление шапки
        ws.freeze_panes = "A2"

        # Лист сводки
        ws_summary = wb.create_sheet(title="Сводка")
        summary = self.tester.get_summary()

        summary_data = [
            ["📊 ОБЩАЯ СТАТИСТИКА", ""],
            ["Всего проверено", summary['total']],
            ["Доступно", summary['accessible']],
            ["Заблокировано", summary['blocked']],
            ["Процент блокировок", f"{summary['block_percentage']:.1f}%"],
            ["", ""],
            ["🔍 ТИПЫ БЛОКИРОВОК", ""],
            ["DNS блокировки", summary['dns_blocks']],
            ["HTTP/TCP блокировки", summary['http_blocks']],
            ["", ""],
            ["💡 РЕКОМЕНДАЦИИ", ""],
        ]

        if summary['block_percentage'] == 0:
            summary_data.append(["Статус", "✅ Блокировок не обнаружено"])
        elif summary['block_percentage'] < 30:
            summary_data.append(["Статус", "⚠️ Частичные блокировки"])
            summary_data.append(["Совет", "Используйте DNS 1.1.1.1 или 8.8.8.8"])
        else:
            summary_data.append(["Статус", "🚫 Значительные блокировки"])
            summary_data.append(["Совет", "Рекомендуется VPN или средства обхода DPI"])

        for row_idx, (label, value) in enumerate(summary_data, 1):
            cell = ws_summary.cell(row=row_idx, column=1, value=label)
            cell.font = Font(bold=True) if "📊" in label or "🔍" in label or "💡" in label else Font()
            ws_summary.cell(row=row_idx, column=2, value=value)

        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 40

        # Сохранение
        wb.save(filename)
        return filename

    def generate_json_report(self) -> str:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{self.report_dir}/report_{timestamp}.json"

        report_data = {
            "timestamp": datetime.now().isoformat(),
            "summary": self.tester.get_summary(),
            "results": []
        }

        for result in self.tester.results:
            report_data["results"].append({
                "resource": {
                    "name": result.resource.name,
                    "domain": result.resource.domain,
                    "category": result.resource.category
                },
                "dns": {
                    "provider_ip": result.dns_result.provider_ip,
                    "public_ip": result.dns_result.public_ip,
                    "is_blocked": result.dns_result.is_blocked,
                    "is_spoofed": result.dns_result.is_spoofed,
                    "response_time": result.dns_result.response_time
                },
                "http": {
                    "is_accessible": result.http_result.is_accessible,
                    "status_code": result.http_result.status_code,
                    "response_time": result.http_result.response_time,
                    "error": result.http_result.error
                },
                "tcp": [
                    {
                        "port": t.port,
                        "is_open": t.is_open,
                        "response_time": t.response_time
                    }
                    for t in result.tcp_results
                ],
                "block_type": result.block_type,
                "confidence": result.confidence,
                "reason": self._get_block_reason(result.block_type),
                "recommendation": self._get_recommendation(result.block_type)
            })

        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(report_data, f, ensure_ascii=False, indent=2)

        return filename

    def generate_html_report(self) -> str:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{self.report_dir}/report_{timestamp}.html"

        summary = self.tester.get_summary()

        html = f"""<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Отчёт о блокировках провайдера</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 40px; background: #f5f5f5; }}
        .container {{ max-width: 1400px; margin: 0 auto; background: white; padding: 30px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
        h1 {{ color: #2c3e50; border-bottom: 3px solid #3498db; padding-bottom: 10px; }}
        .summary {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 20px; margin: 30px 0; }}
        .stat-card {{ background: #ecf0f1; padding: 20px; border-radius: 8px; text-align: center; }}
        .stat-value {{ font-size: 2em; font-weight: bold; color: #2c3e50; }}
        .stat-label {{ color: #7f8c8d; margin-top: 5px; }}
        table {{ width: 100%; border-collapse: collapse; margin: 20px 0; font-size: 0.9em; }}
        th, td {{ padding: 12px; text-align: left; border-bottom: 1px solid #ddd; }}
        th {{ background: #3498db; color: white; }}
        tr:hover {{ background: #f5f5f5; }}
        .status-ok {{ color: #27ae60; font-weight: bold; }}
        .status-blocked {{ color: #e74c3c; font-weight: bold; }}
        .timestamp {{ color: #7f8c8d; font-size: 0.9em; }}
        .recommendation {{ background: #fff3cd; padding: 15px; border-radius: 5px; margin: 20px 0; }}
    </style>
</head>
<body>
    <div class="container">
        <h1>🛡️ Отчёт о блокировках провайдера</h1>
        <p class="timestamp">Дата проверки: {datetime.now().strftime("%d.%m.%Y %H:%M")}</p>

        <div class="summary">
            <div class="stat-card">
                <div class="stat-value">{summary['total']}</div>
                <div class="stat-label">Всего проверено</div>
            </div>
            <div class="stat-card">
                <div class="stat-value" style="color: #27ae60;">{summary['accessible']}</div>
                <div class="stat-label">Доступно</div>
            </div>
            <div class="stat-card">
                <div class="stat-value" style="color: #e74c3c;">{summary['blocked']}</div>
                <div class="stat-label">Заблокировано</div>
            </div>
            <div class="stat-card">
                <div class="stat-value" style="color: #f39c12;">{summary['block_percentage']:.1f}%</div>
                <div class="stat-label">Процент блокировок</div>
            </div>
        </div>

        <h2>📋 Детальный отчёт</h2>
        <table>
            <tr>
                <th>Ресурс</th>
                <th>Домен</th>
                <th>Статус</th>
                <th>Тип блокировки</th>
                <th>Порт 80</th>
                <th>Порт 443</th>
                <th>Рекомендация</th>
            </tr>
"""

        for result in self.tester.results:
            status_class = "status-ok" if result.block_type == "none" else "status-blocked"
            status_text = "✅ Доступен" if result.block_type == "none" else "🚫 Заблокирован"

            port_80 = "✅" if any(t.port == 80 and t.is_open for t in result.tcp_results) else "❌"
            port_443 = "✅" if any(t.port == 443 and t.is_open for t in result.tcp_results) else "❌"

            html += f"""            <tr>
                <td>{result.resource.name}</td>
                <td>{result.resource.domain}</td>
                <td class="{status_class}">{status_text}</td>
                <td>{result.block_type}</td>
                <td>{port_80}</td>
                <td>{port_443}</td>
                <td>{self._get_recommendation(result.block_type)}</td>
            </tr>
"""

        html += """        </table>

        <div class="recommendation">
            <h3>💡 Общие рекомендации</h3>
"""

        if summary['block_percentage'] == 0:
            html += "<p style='color: #27ae60;'>✅ Ваш провайдер не блокирует проверенные ресурсы!</p>"
        else:
            html += """            <ul>
                <li>Используйте DNS 1.1.1.1 или 8.8.8.8</li>
                <li>Для обхода DPI: GoodbyeDPI, Zapret</li>
                <li>Для полного доступа: VPN или прокси</li>
            </ul>
"""

        html += """        </div>
    </div>
</body>
</html>
"""

        with open(filename, 'w', encoding='utf-8') as f:
            f.write(html)

        return filename


# ============================================================================
# ГЛАВНАЯ ФУНКЦИЯ
# ============================================================================

def print_banner():
    banner = f"""
{Fore.CYAN}
╔═══════════════════════════════════════════════════════════╗
║                                                           ║
║        🛡️  PROVIDER BLOCK TESTER  v3.1                    ║
║                                                           ║
║     Тестер блокировок с Excel отчётом на рабочем столе    ║
║                                                           ║
╚═══════════════════════════════════════════════════════════╝
{Style.RESET_ALL}
    """
    print(banner)


def main():
    parser = argparse.ArgumentParser(
        description='Тестер блокировок провайдера',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Примеры использования:
  python provider_tester.py                    # Полный тест
  python provider_tester.py --excel            # + Excel отчёт на рабочий стол
  python provider_tester.py --all              # Все форматы (Excel + JSON + HTML)
  python provider_tester.py --quick            # Быстрый тест
  python provider_tester.py --resource youtube.com  # Тест одного ресурса
        """
    )

    parser.add_argument('--excel', action='store_true', help='Сохранить Excel отчёт на рабочий стол')
    parser.add_argument('--json', action='store_true', help='Сохранить JSON отчёт')
    parser.add_argument('--html', action='store_true', help='Сохранить HTML отчёт')
    parser.add_argument('--all', action='store_true', help='Все форматы отчётов')
    parser.add_argument('--quick', action='store_true', help='Быстрый тест')
    parser.add_argument('--resource', type=str, help='Тестировать конкретный ресурс')
    parser.add_argument('--verbose', '-v', action='store_true', help='Подробный вывод')
    parser.add_argument('--no-pause', action='store_true', help='Не ждать нажатия клавиши в конце')

    args = parser.parse_args()

    print_banner()

    print(f"{Fore.YELLOW}⚙️  Инициализация тестера...{Style.RESET_ALL}")
    tester = ProviderTester(DEFAULT_CONFIG)

    if args.quick:
        print(f"{Fore.YELLOW}🚀 Быстрый режим активирован{Style.RESET_ALL}")
        DEFAULT_CONFIG.TEST_PORTS = [443]

    if args.resource:
        print(f"{Fore.YELLOW}🎯 Тестирование конкретного ресурса: {args.resource}{Style.RESET_ALL}")
        DEFAULT_CONFIG.TEST_RESOURCES = [
            TestResource(args.resource, args.resource, f"https://{args.resource}", "custom")
        ]

    try:
        results = tester.run_all_tests()
    except KeyboardInterrupt:
        print(f"\n{Fore.RED}❌ Тестирование прервано пользователем{Style.RESET_ALL}")
        input("\nНажмите Enter для выхода...")
        sys.exit(1)
    except Exception as e:
        print(f"\n{Fore.RED}❌ Ошибка при тестировании: {e}{Style.RESET_ALL}")
        if args.verbose:
            import traceback
            traceback.print_exc()
        input("\nНажмите Enter для выхода...")
        sys.exit(1)

    reporter = ReportGenerator(tester)
    reporter.generate_console_report()

    reports_saved = []

    # Excel всегда создаётся если указан флаг --excel или --all
    if args.excel or args.all:
        if EXCEL_AVAILABLE:
            try:
                excel_file = reporter.generate_excel_report(on_desktop=True)  # На рабочий стол
                if excel_file:
                    reports_saved.append(excel_file)
                    print(f"{Fore.GREEN}📊 Excel отчёт сохранён на рабочий стол: {excel_file}{Style.RESET_ALL}")
            except Exception as e:
                print(f"{Fore.RED}❌ Ошибка сохранения Excel: {e}{Style.RESET_ALL}")
        else:
            print(f"{Fore.RED}❌ Excel отчёт недоступен (установите openpyxl){Style.RESET_ALL}")

    if args.json or args.all:
        try:
            json_file = reporter.generate_json_report()
            reports_saved.append(json_file)
            print(f"{Fore.GREEN}📄 JSON отчёт сохранён: {json_file}{Style.RESET_ALL}")
        except Exception as e:
            print(f"{Fore.RED}❌ Ошибка сохранения JSON: {e}{Style.RESET_ALL}")

    if args.html or args.all:
        try:
            html_file = reporter.generate_html_report()
            reports_saved.append(html_file)
            print(f"{Fore.GREEN}📄 HTML отчёт сохранён: {html_file}{Style.RESET_ALL}")
        except Exception as e:
            print(f"{Fore.RED}❌ Ошибка сохранения HTML: {e}{Style.RESET_ALL}")

    if reports_saved:
        print(f"\n{Fore.CYAN}📁 Отчёты сохранены{Style.RESET_ALL}")

    # ФИНАЛЬНОЕ СООБЩЕНИЕ
    print(f"\n{Fore.CYAN}{'=' * 70}{Style.RESET_ALL}")
    print(f"{Fore.GREEN}✅ ТЕСТИРОВАНИЕ ЗАВЕРШЕНО!{Style.RESET_ALL}")
    print(f"{Fore.CYAN}{'=' * 70}{Style.RESET_ALL}")

    # Пауза перед закрытием (если не указан флаг --no-pause)
    if not args.no_pause:
        print(f"\n{Fore.YELLOW}⏸️  Нажмите Enter для выхода...{Style.RESET_ALL}")
        input()

    return 0


if __name__ == "__main__":
    sys.exit(main())